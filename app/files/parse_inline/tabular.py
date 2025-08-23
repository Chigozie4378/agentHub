import csv
from pathlib import Path

def extract_csv_or_xlsx(path: str, max_rows: int = 50) -> tuple[str, dict]:
    p = Path(path)
    if p.suffix.lower() == ".csv":
        return _extract_csv(path, max_rows)
    if p.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl  # optional
        except Exception:
            return ("[XLSX detected; install openpyxl to preview rows]", {"rows": 0, "cols": 0})
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = []
        cols = ws.max_column
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(",".join("" if v is None else str(v) for v in row))
            if i + 1 >= max_rows:
                break
        return ("\n".join(rows)), {"rows": min(ws.max_row, max_rows), "cols": cols}
    return ("", {"rows": 0, "cols": 0})

def _extract_csv(path: str, max_rows: int) -> tuple[str, dict]:
    out = []
    cols = 0
    with open(path, newline="", encoding="utf-8", errors="ignore") as f:
        r = csv.reader(f)
        for i, row in enumerate(r):
            cols = max(cols, len(row))
            out.append(",".join(row))
            if i + 1 >= max_rows:
                break
    return ("\n".join(out)), {"rows": len(out), "cols": cols}
